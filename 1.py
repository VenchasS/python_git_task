import csv
import re
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

from unittest import TestCase, main



currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}
thin = Side(border_style="thin", color="000000")


wb = Workbook()
ws1 = wb.create_sheet("Статистика по годам", 0)
ws2 = wb.create_sheet("Статистика по городам", 1)



def generate_excel():
    """ Создает exel обьект
    """
    wb = Workbook()
    ws1 = wb.create_sheet("Статистика по годам", 0)
    ws2 = wb.create_sheet("Статистика по городам", 1)
    wb.remove(wb['Sheet'])
    wb.save("report.xlsx")

def clean(text):
    """
    очищает строку
    :param text: исходная строка
    :return: очищенная строка
    """
    example = re.compile(r'<[^>]+>')
    s = example.sub('', text).replace(' ', ' ').replace('\xa0', ' ').strip()
    return re.sub(" +", " ", s)

def csv_reader(name):
    """

    :param name: имя исходного файла
    :return: возвращает, заголовки и данные
    """
    csv_list = csv.reader(open(name, encoding='utf-8-sig'))
    data = [x for x in csv_list]
    return data[0], data[1::]

def csv_filer(reader):
    """
    Проверяет данные
    :param reader: ридер csv файла
    :return: возвращает структурированные данные
    """
    all_vac = [x for x in reader if '' not in x and len(x) == len(reader[0])]
    vac = [[clean(y) for y in x] for x in all_vac]
    return vac

name = input('Введите название файла: ')
if name == "":
    name = "vacancies_by_year.csv"
prof = input('Введите название профессии: ')
header, vac = csv_reader(name)
vac = csv_filer(vac)
dict_naming = {}
for i in range(len(header)):
    dict_naming[header[i]] = i

salary_dynamic = {}
count_dynamic = {}
salary_prof_dynamic = {}
city_count = {}
salary_city = {}
prof_count = {}
for item in vac:
    year = int(item[dict_naming['published_at']].split('-')[0])
    if year not in count_dynamic:
        count_dynamic[year] = 0
    count_dynamic[year] += 1
    for i in range(len(item)):
        if header[i] == 'salary_from':
            salary = (float(item[i]) + float(item[i+1])) / 2
            if item[dict_naming['salary_currency']] != 'RUR':
                salary *= currency_to_rub[item[dict_naming['salary_currency']]]
            if year not in salary_dynamic:
                salary_dynamic[year] = []
            salary_dynamic[year].append(int(salary))
            if year not in salary_prof_dynamic:
                salary_prof_dynamic[year] = []
            if prof in item[0]: salary_prof_dynamic[year].append(int(salary))
            if year not in prof_count:
                prof_count[year] = 0
            if prof in item[0]: prof_count[year] += 1
        city = item[dict_naming['area_name']]
        if city not in city_count:
            city_count[city] = 0
        city_count[city] += 1

for item in vac:
    for i in range(len(item)):
        if header[i] == 'salary_from':
            salary = (float(item[i]) + float(item[i+1])) / 2
            city = item[dict_naming['area_name']]
            if item[dict_naming['salary_currency']] != 'RUR':
                salary *= currency_to_rub[item[dict_naming['salary_currency']]]
            if city_count[city] >= int(sum(city_count.values()) * 0.01):
                if city not in salary_city:
                    salary_city[city] = []
                salary_city[city].append(int(salary))

for key in salary_dynamic:
    salary_dynamic[key] = sum(salary_dynamic[key]) // len(salary_dynamic[key])

for key in salary_prof_dynamic:
    salary_prof_dynamic[key] = sum(salary_prof_dynamic[key]) // max(len(salary_prof_dynamic[key]), 1)

for key in salary_city:
    salary_city[key] = sum(salary_city[key]) // len(salary_city[key])

def set_cities(salaries, vacancies):
    """
    Устанавливает в таблицу все списки городов
    :param salaries: список зарплат
    :param vacancies: список вакансий
    :return: ничего
    """
    setCell(row=1, column=1, value="Город", ws=ws2,bold=True)
    setCell(row=1, column=2, value="Уровень зарплаты", ws=ws2,bold=True)
    setCell(row=1, column=3, value="Город", ws=ws2,bold=True)
    setCell(row=1, column=4, value="Доля вакансий", ws=ws2,bold=True)
    index = 2
    for key in salaries:
        setCell(row=index, column=1, value=key, ws=ws2)
        setCell(row=index, column=2, value=salaries[key], ws=ws2)
        index+=1
    index = 2
    for key in vacancies:
        setCell(row=index, column=3, value=key, ws=ws2)
        setCell(row=index, column=4, value=vacancies[key], ws=ws2)
        index+=1

def setCell(row,column,value,ws, bold=False):
    """
    Устанавливает значение в таблицу применяя параметры
    :param row: номер строки
    :param column   номер столбца
    :param value: значение ячейки
    :param ws: ссылка на таблицу
    :param bold: жирность текста
    :return: ничего
    """
    cell = ws.cell(row=row, column=column,value=value)
    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    if bold:
        cell.font = Font(bold=True)


def set_years(salary_dynamic, count_dynamic, salary_prof_dynamic , prof_count):
    """
    Устанавливает значение в таблицу применяя параметры
    :param salary_dynamic: список зарплат
    :param count_dynamic: список кол-ва
    :param salary_prof_dynamic: список зарплат по професии
    :param prof_count: список кол-ва по профессии
    :return: ничего
    """
    setCell(row=1,column=1,value="год",ws=ws1,bold=True)
    setCell(row=1,column=2,value="Средняя зарплата" ,ws=ws1,bold=True)
    setCell(row=1,column=3,value="Средняя зарплата " + prof ,ws=ws1,bold=True)
    setCell(row=1,column=4,value="Количество вакансий",ws=ws1,bold=True)
    setCell(row=1,column=5,value="Количество вакансий " + prof,ws=ws1,bold=True)
    index = 2
    for key in prof_count:
        setCell(row=index, column=1, value=key, ws=ws1)
        setCell(row=index, column=2, value=salary_dynamic[key], ws=ws1)
        setCell(row=index, column=3, value=salary_prof_dynamic[key], ws=ws1)
        setCell(row=index, column=4, value=count_dynamic[key], ws=ws1)
        setCell(row=index, column=5, value=prof_count[key], ws=ws1)
        index+=1





print('Динамика уровня зарплат по годам:', salary_dynamic)
print('Динамика количества вакансий по годам:', count_dynamic)
print('Динамика уровня зарплат по годам для выбранной профессии:', salary_prof_dynamic)
print('Динамика количества вакансий по годам для выбранной профессии:', prof_count)
set_years(salary_dynamic=salary_dynamic,count_dynamic=count_dynamic,salary_prof_dynamic=salary_prof_dynamic,prof_count=prof_count)

print('Уровень зарплат по городам (в порядке убывания):', dict(Counter(salary_city).most_common(10)))
most = dict(Counter({k: float('{:.4f}'.format(v/sum(city_count.values()))) for k,v in city_count.items()}).most_common(10))
most = {k: v for k, v in most.items() if v >= 0.01}
print('Доля вакансий по городам (в порядке убывания):', most)
set_cities(salaries=dict(Counter(salary_city).most_common(10)), vacancies=most)
generate_excel()

class UnitTest(TestCase):

    def test_clean(self):
        self.assertEqual(clean("<p><strong>ОБЯЗАННОСТИ:</strong></p>"), "ОБЯЗАННОСТИ:")
        self.assertEqual(clean("<p><strong>Cotvec </strong>- IT-компания, которая занимается консалтингом и разработкой программного обеспечения для организаций банковско-финансового сектора.</p>"),"Cotvec - IT-компания, которая занимается консалтингом и разработкой программного обеспечения для организаций банковско-финансового сектора.")

    def test_csv_filer(self):
        self.assertEqual(csv_filer([[
            "test"
        ]])[0][0], "test")

    def test_setCell(self):
        generate_excel()
        row = 1
        column = 1
        value = 2289
        setCell(row,column,value,ws1)
        sell = ws1.cell(row=row, column=column)
        self.assertEqual(value, sell.value)

    def test_csv_reader(self):
        self.assertEqual(header[0],"name")
        self.assertEqual(header[1], "salary_from")
        self.assertEqual(header[2], "salary_to")
        pass

if __name__ == '__main__':
    main()